import os
from openpyxl import load_workbook, Workbook

def prepare_sheet_in_memory(file_path):
    # Load the workbook and select the active worksheet
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    # Extract file name for use as header
    file_name = os.path.basename(file_path)

    # Insert a new row at the top and add 'file name' heading and file name
    # We do this in memory and do not save the file
    sheet.insert_rows(1)
    sheet['A1'] = 'file name'
    sheet['B1'] = file_name
    
    return sheet

def sheet_to_dict(sheet):
    # Skip the file name row and convert sheet data to a dictionary
    data_dict = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        key = row[0]
        values = row[1:]
        data_dict[key] = values
    return data_dict

def merge_dicts(dict1, dict2):
    # Merge dict2 into dict1 based on keys, combining values for matching keys
    for key in dict2:
        if key in dict1:
            dict1[key] = dict1[key] + dict2[key]
        else:
            dict1[key] = dict2[key]
    return dict1

def main():
    # Paths to the original files
    file1_path = r"C:\Lectrix_company\work\Git_Projects\Automationdashboard\Automationdashboard\analysis_B4_BT02_09.56_10.40.xlsx"
    file2_path = r"C:\Lectrix_company\work\Git_Projects\Automationdashboard\Automationdashboard\analysis_B4_BT02_13.59_14.50.xlsx"
    
    # Prepare both files in memory
    sheet1 = prepare_sheet_in_memory(file1_path)
    sheet2 = prepare_sheet_in_memory(file2_path)

    # Convert both sheets to dictionaries
    dict1 = sheet_to_dict(sheet1)
    dict2 = sheet_to_dict(sheet2)

    # Merge the dictionaries
    merged_data = merge_dicts(dict1, dict2)

    # Create a new workbook for the merged data
    merged_workbook = Workbook()
    merged_sheet = merged_workbook.active

    # Add customized headers based on extracted file names
    file1_name = os.path.basename(file1_path)
    file2_name = os.path.basename(file2_path)
    merged_sheet.append(['File name', file1_name, file2_name])

    for key, values in merged_data.items():
        merged_sheet.append([key] + list(values))

    # Define a path for the merged workbook to be saved
    merged_file_path = r'C:\Lectrix_company\work\Git_Projects\Automationdashboard\Automationdashboard\merged_file_source_directory.xlsx'
    # Save the merged workbook
    merged_workbook.save(filename=merged_file_path)

if __name__ == '__main__':
    main()
